#@title Célula 2 · Lógica do sistema { display-mode: "form" }
#@markdown Execute uma única vez por sessão.
#@markdown
# **sync: AAAA-MM-DD** — atualize a data ao colar uma nova versão do `logic.py`.
#
# ---
#
# _Para adicionar um fornecedor: crie `parse_fatura_<nome>` e registre em `PARSERS`. Veja o README._

import io
import pandas as pd
import openpyxl


# ── Constantes ────────────────────────────────────────────────────────────────

TOLERANCIA = 0.05  # diferença máxima em R$ para considerar OK (arredondamento)

REFERENCIA_KEYWORDS = ("descontos", "proventos")
# Substrings que devem estar presentes no nome do arquivo de referência interna.


# ════════════════════════════════════════════════════════════════════════════════
# PARSERS DE FATURA
#
# Contrato: toda função parse_fatura_<nome> deve
#   · receber conteudo: bytes
#   · retornar DataFrame com colunas:
#       matricula       int    matrícula do colaborador
#       desconto_fatura float  desconto total do colaborador (soma de todas as vidas)
#       custo_fatura    float  custo total para a empresa
#       qtd_vidas       int    número de vidas cobertas (titular + dependentes)
#   · agregar por matricula quando o arquivo tiver uma linha por vida
#   · lançar exceção com mensagem descritiva em caso de estrutura inesperada
# ════════════════════════════════════════════════════════════════════════════════

def parse_fatura_bradesco_dental(conteudo):
    """
    Parser da fatura mensal do Bradesco Dental.

    Estrutura esperada:
      · Aba: bradesco
      · Cabeçalho localizado dinamicamente pela presença de 'Matricula Titular'
      · Uma linha por vida (titular e dependentes)
      · Colunas lidas: Matricula Titular, DESCONTO COLABORADOR, custo (coluna X)
      · Agrega por matrícula antes de retornar
    """
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb["bradesco"]

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(str(v).strip().lower() == "matricula titular" for v in row if v):
            header_row = i
            break
    if header_row is None:
        raise ValueError("Coluna 'Matricula Titular' não encontrada na aba 'bradesco'.")

    headers = list(ws.iter_rows(values_only=True))[header_row]
    idx = {}
    for i, h in enumerate(headers):
        h_norm = str(h).strip().lower() if h else ""
        if h_norm == "matricula titular":
            idx["mat"] = i
        elif h_norm == "desconto colaborador":
            idx["desconto"] = i
        elif h_norm in ["x", "custo"] and "custo" not in idx:
            idx["custo"] = i

    registros = []
    for row in list(ws.iter_rows(values_only=True))[header_row + 2:]:
        mat   = row[idx["mat"]]      if "mat"     in idx else None
        desc  = row[idx["desconto"]] if "desconto" in idx else None
        custo = row[idx["custo"]]    if "custo"   in idx else None
        if mat and isinstance(mat, (int, float)) and desc:
            registros.append({
                "matricula":       int(mat),
                "desconto_fatura": float(desc),
                "custo_fatura":    float(custo) if custo else 0.0,
            })

    df = pd.DataFrame(registros)
    return df.groupby("matricula").agg(
        desconto_fatura=("desconto_fatura", "sum"),
        custo_fatura=("custo_fatura", "sum"),
        qtd_vidas=("matricula", "count"),
    ).reset_index()


def parse_fatura_unimed(conteudo):
    """
    Parser da fatura mensal da Unimed.

    Estrutura esperada:
      · Aba: unimed
      · Cabeçalho localizado dinamicamente pela presença de 'Matricula Titular'
      · Uma linha por vida (titular e dependentes)
      · Colunas lidas: Matricula Titular, DESCONTO COLABORADOR, CUSTO ATLANTICO
      · Agrega por matrícula antes de retornar
    """
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb["unimed"]

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(str(v).strip().lower() == "matricula titular" for v in row if v):
            header_row = i
            break
    if header_row is None:
        raise ValueError("Coluna 'Matricula Titular' não encontrada na aba 'unimed'.")

    headers = list(ws.iter_rows(values_only=True))[header_row]
    idx = {}
    for i, h in enumerate(headers):
        h_norm = str(h).strip().lower() if h else ""
        if h_norm == "matricula titular":
            idx["mat"] = i
        elif h_norm == "desconto colaborador":
            idx["desconto"] = i
        elif h_norm == "custo atlantico":
            idx["custo"] = i

    missing = [k for k in ("mat", "desconto", "custo") if k not in idx]
    if missing:
        raise ValueError(f"Colunas não encontradas na aba 'unimed': {missing}")

    registros = []
    for row in list(ws.iter_rows(values_only=True))[header_row + 1:]:
        mat   = row[idx["mat"]]
        desc  = row[idx["desconto"]]
        custo = row[idx["custo"]]
        if mat and isinstance(mat, (int, float)) and isinstance(desc, (int, float)):
            registros.append({
                "matricula":       int(mat),
                "desconto_fatura": float(desc),
                "custo_fatura":    float(custo) if isinstance(custo, (int, float)) else 0.0,
            })

    df = pd.DataFrame(registros)
    return df.groupby("matricula").agg(
        desconto_fatura=("desconto_fatura", "sum"),
        custo_fatura=("custo_fatura", "sum"),
        qtd_vidas=("matricula", "count"),
    ).reset_index()


# ── Registro de fornecedores ──────────────────────────────────────────────────
#
# Chave: tupla de substrings que devem estar presentes no nome do arquivo
#        (comparação em lowercase, todas devem estar presentes)
# Valor: função de parse correspondente
#
PARSERS = {
    ("bradesco", "dental"): parse_fatura_bradesco_dental,
    ("unimed",):            parse_fatura_unimed,
    # ("uniodonto",):        parse_fatura_uniodonto,
}

# Mapeamento: chave do parser → label do fornecedor na planilha de referência.
# Deve bater com o texto do cabeçalho da aba 'total' (case-insensitive).
# Ao adicionar um fornecedor em PARSERS, adicione o label correspondente aqui.
FORNECEDOR_LABELS = {
    ("bradesco", "dental"): "bradesco dental",
    ("unimed",):            "unimed",
    # ("uniodonto",):        "uniodonto",
}


# ════════════════════════════════════════════════════════════════════════════════
# PARSER DE REFERÊNCIA INTERNA
# ════════════════════════════════════════════════════════════════════════════════

def parse_referencia_interna(conteudo, fornecedor_label):
    """
    Parser da referência interna do DP.

    Parâmetros
    ----------
    conteudo : bytes
        Conteúdo binário do arquivo xlsx.
    fornecedor_label : str
        Label do fornecedor conforme aparece no cabeçalho da planilha
        (ex: 'bradesco dental', 'unimed'). Case-insensitive.

    Estrutura esperada:
      · Arquivo: descontos_proventos_beneficios_MMAAAA.xlsx
      · Aba: total
      · Cabeçalho em duas linhas:
          Linha 1: Mat, Nome, ..., <fornecedor_label>, ...
          Linha 2: subcolunas Fatura, Desconto, Custo
      · Busca de Fatura/Desconto/Custo limitada ao bloco do fornecedor
        (até 10 colunas a partir da coluna do fornecedor)
    """
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb["total"]
    rows = list(ws.iter_rows(values_only=True))

    idx_mat = idx_nome = idx_fatura = idx_desconto = idx_custo = None
    idx_bd_start = header1_row = header2_row = None

    for i, row in enumerate(rows):
        vals = [str(v).strip().lower() if v else "" for v in row]
        if "mat" in vals and "nome" in vals:
            header1_row = i
            idx_mat  = next(j for j, v in enumerate(vals) if v == "mat")
            idx_nome = next(j for j, v in enumerate(vals) if v == "nome")
            for j, v in enumerate(row):
                if v and fornecedor_label.lower() in str(v).lower():
                    idx_bd_start = j
                    header2_row  = i + 1
                    break
            break

    if header2_row is None or idx_bd_start is None:
        raise ValueError(
            f"Cabeçalho do fornecedor '{fornecedor_label}' não encontrado "
            f"na aba 'total'. Verifique se o nome bate com o cabeçalho da planilha."
        )

    subheaders = rows[header2_row]
    for j in range(idx_bd_start, min(idx_bd_start + 10, len(subheaders))):
        v = subheaders[j]
        if v is None:
            continue
        v_low = str(v).strip().lower()
        if "fatura" in v_low and idx_fatura is None:
            idx_fatura = j
        elif "desconto" in v_low and idx_fatura is not None and idx_desconto is None:
            idx_desconto = j
        elif "custo" in v_low and idx_desconto is not None and idx_custo is None:
            idx_custo = j
            break

    if None in (idx_fatura, idx_desconto, idx_custo):
        raise ValueError(
            f"Subcolunas Fatura/Desconto/Custo não encontradas para '{fornecedor_label}'. "
            f"Índices encontrados: fatura={idx_fatura}, desconto={idx_desconto}, custo={idx_custo}."
        )

    data_start = header2_row + 1
    registros  = []
    for row in rows[data_start:]:
        mat   = row[idx_mat]      if idx_mat      is not None else None
        nome  = row[idx_nome]     if idx_nome     is not None else None
        fat   = row[idx_fatura]   if idx_fatura   is not None else None
        desc  = row[idx_desconto] if idx_desconto is not None else None
        custo = row[idx_custo]    if idx_custo    is not None else None
        if mat and isinstance(mat, (int, float)):
            registros.append({
                "matricula":         int(mat),
                "nome":              nome,
                "desconto_esperado": float(desc)  if isinstance(desc,  (int, float)) else 0.0,
                "fatura_esperada":   float(fat)   if isinstance(fat,   (int, float)) else 0.0,
                "custo_esperado":    float(custo) if isinstance(custo, (int, float)) else 0.0,
            })
    return pd.DataFrame(registros)


# ════════════════════════════════════════════════════════════════════════════════
# NÚCLEO DE CONCILIAÇÃO
# ════════════════════════════════════════════════════════════════════════════════

def conciliar(df_fatura, df_referencia):
    """
    Cruza fatura e referência interna por matrícula (join outer).

    Retorna DataFrame com as colunas originais de ambos os lados mais:
      diferenca   float   desconto_fatura - desconto_esperado
      status      str     classificação da linha (ver tabela abaixo)

    Classificação de status:
      ✅ OK                  |diferenca| <= TOLERANCIA
      👻 Só na fatura        presente na fatura, ausente na referência
      🔍 Só na referência    presente na referência, ausente na fatura
      💰 Divergência de valor presente nos dois lados, diferença fora da tolerância
    """
    df = df_referencia.merge(df_fatura, on="matricula", how="outer")
    df["desconto_fatura"]   = df["desconto_fatura"].fillna(0.0)
    df["desconto_esperado"] = df["desconto_esperado"].fillna(0.0)
    df["diferenca"]         = (df["desconto_fatura"] - df["desconto_esperado"]).round(4)

    def _status(row):
        if abs(row["diferenca"]) <= TOLERANCIA:  return "✅ OK"
        elif row["desconto_esperado"] == 0:       return "👻 Só na fatura"
        elif row["desconto_fatura"]   == 0:       return "🔍 Só na referência"
        else:                                     return "💰 Divergência de valor"

    df["status"] = df.apply(_status, axis=1)
    return df


# ── Utilitários ───────────────────────────────────────────────────────────────

def identificar_fornecedor(nome_arquivo):
    """
    Retorna (chave, fn_parser) para o fornecedor identificado pelo nome do arquivo,
    ou (None, None) se nenhum parser for encontrado.
    """
    nome_lower = nome_arquivo.lower()
    for chave, fn in PARSERS.items():
        if all(k in nome_lower for k in chave):
            return chave, fn
    return None, None


fornecedores_disponiveis = [
    " ".join(chave).title() for chave in PARSERS.keys()
]